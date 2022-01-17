''' This script file is used to capture interface status, BGP and BFD status in excel file for pre and post. After storing output in excel also compare interface status,
BGP state, BGP prefixes and BFD status with precheck and provide us result as a changed or not. '''

import getpass
import re

import openpyxl

import openpyxl
import openpyxl.utils
import openpyxl.formatting
import openpyxl.formula
import openpyxl.styles
from openpyxl.styles.differential import DifferentialStyle
import openpyxl.formatting.rule
import netmiko


device_list = []  # Store device list for which pre/post perform
row = 0
col = 0

no_device = int(input("Please enter number of devices  "))
for i in range(no_device):
    device_list.append(input("Please enter device number {}   ".format(i+1)))

# Globally define style for cell headings and cells
heading_style_template_precheck = openpyxl.styles.NamedStyle(name="heading_style_template_precheck",
                                                             font=openpyxl.styles.Font(size=12, bold=True),
                                                             border=openpyxl.styles.Border(left=openpyxl.styles.Side(style="thin", color="00000000"),
                                                                                  right=openpyxl.styles.Side(style="thin", color="00000000"),
                                                                                  top=openpyxl.styles.Side(style="thin", color="00000000"),
                                                                                  bottom=openpyxl.styles.Side(style="thin", color="00000000")),
                                                             fill=openpyxl.styles.PatternFill(patternType="solid", fgColor="00FF6600", bgColor="00CCFFFF"))
heading_style_template_postcheck = openpyxl.styles.NamedStyle(name="heading_style_template_postcheck",
                                                              font=openpyxl.styles.Font(size=12, bold=True),
                                                              border=openpyxl.styles.Border(left=openpyxl.styles.Side(style="thin", color="00000000"),
                                                                                  right=openpyxl.styles.Side(style="thin", color="00000000"),
                                                                                  top=openpyxl.styles.Side(style="thin", color="00000000"),
                                                                                  bottom=openpyxl.styles.Side(style="thin", color="00000000")),
                                                              fill=openpyxl.styles.PatternFill(patternType="solid", fgColor="00FF6600", bgColor="00CCFFFF"))
cell_style_template_precheck = openpyxl.styles.NamedStyle(name="cell_style_template_precheck",
                                                          border=openpyxl.styles.Border(left=openpyxl.styles.Side(style="thin", color="00000000"),
                                                                                  right=openpyxl.styles.Side(style="thin", color="00000000"),
                                                                                  top=openpyxl.styles.Side(style="thin", color="00000000"),
                                                                                  bottom=openpyxl.styles.Side(style="thin", color="00000000")),
                                                          alignment=openpyxl.styles.Alignment(horizontal="center"))
cell_style_template_postcheck = openpyxl.styles.NamedStyle(name="cell_style_template_postcheck",
                                                 border=openpyxl.styles.Border(left=openpyxl.styles.Side(style="thin", color="00000000"),
                                                                                  right=openpyxl.styles.Side(style="thin", color="00000000"),
                                                                                  top=openpyxl.styles.Side(style="thin", color="00000000"),
                                                                                  bottom=openpyxl.styles.Side(style="thin", color="00000000")),
                                                 alignment=openpyxl.styles.Alignment(horizontal="center"))
cell_style_matched_condition = openpyxl.styles.NamedStyle(name="cell_style_matched_condition",
                                                          fill=openpyxl.styles.PatternFill(patternType="solid", bgColor="00FF0000"))


username = input("Please enter username ")
password = getpass.getpass("Please enter password ")
purpose = input("Please enter 'pre' for pre-check and 'post' for post-check logs    ") # To check is it for pre or post

if purpose.lower() == 'pre':
    try:
        with open("main block error.txt", mode="w") as main_error:  # To write if error occured
                for i in device_list:
                    cleaned_int_status = []  # To store data after cleaning
                    cleaned_bgp_summary = []
                    bgp_neighbor = []  # used to store output of device
                    bgp_found_list = []  # Used to store all list if found BGP in list
                    memory_found_list = []  # Same way store memory
                    bgp_neighbor_state = []  # BGP neighbor state only
                    bgp_neighbor_state_splitted = []  # BGP neighbor state only, but it will have new line and spaces.

                    # Login into devices and capture required commands output
                    device_ssh = netmiko.ConnectHandler(host=str(i), username=username,
                                                       password=password,
                                                       device_type="cisco_nxos")

                    device_ssh.find_prompt()
                    print(device_ssh.find_prompt())
                    interface_status = device_ssh.send_command("show interface status")
                    bgp_summary = device_ssh.send_command("show ip bgp vrf all summary")
                    bfd_neighbor = device_ssh.send_command("show bfd neighbor")
                    # Write output in text files for further processing
                    with open(r"Device_Output_Files\int_status_{}.txt".format(i), mode="w") as text_file_int:
                        text_file_int.write(interface_status)
                        text_file_int.close()
                    with open(r"Device_Output_Files\bgp_summary_{}.txt".format(i), mode="w") as text_file_bgp:
                        text_file_bgp.write(bgp_summary)
                        text_file_bgp.close()
                    with open(r"Device_Output_Files\bfd_neighbor_{}.txt".format(i), mode="w") as text_file_bfd:
                        text_file_bfd.write(bfd_neighbor)
                        text_file_bfd.close()
                    device_ssh.disconnect()

                    # Open device output files in readonly mode
                    with open(r"Device_Output_Files\int_status_{}.txt".format(i)) as text_file2_int:
                        new_interface_status = text_file2_int.readlines()
                    with open(r"Device_Output_Files\bgp_summary_{}.txt".format(i)) as text_file2_bgp:
                        bgp_summary_output = text_file2_bgp.readlines()
                    with open(r"Device_Output_Files\bfd_neighbor_{}.txt".format(i)) as text_file2_bfd:
                        bfd_neighbor_output = text_file2_bfd.readlines()

                    # Open workbook or Excel files to write data
                    wb_pre = openpyxl.Workbook()
                    precheck_sheet1 = wb_pre.active
                    precheck_sheet1.title = "Interfaces_precheck"
                    precheck_sheet2 = wb_pre.create_sheet(title="BGP_Summary_Precheck")
                    precheck_sheet3 = wb_pre.create_sheet(title="BFD_neighbor_precheck")

                    # Clean interface status to record in Excel file which received from device_output_int.txt
                    re_pattern_all_aplhabet_start = re.compile(r"(^([a-z]|[A-Z])+)")  # Define pattern to match interfaces
                    for k in range(len(new_interface_status)):
                        re_match_all_alphabet_start = re_pattern_all_aplhabet_start.finditer(new_interface_status[k])
                        if re_match_all_alphabet_start:  # If match found then splitted and appended in list
                            for match in re_match_all_alphabet_start:
                                cleaned_int_status.append(new_interface_status[k].rstrip("\n").split(" "))

                    # Each elements which found with match that is string and after splitting string become list
                    # So iterate in that list and add values in Excel sheet
                    row_int = 0  # Reset this value so that previous loop value will discard out
                    for element_index in range(len(cleaned_int_status)):
                        row_int += 1
                        col_int = 1
                        for r in range(len(cleaned_int_status[element_index])):
                            if cleaned_int_status[element_index][r] != "":
                                precheck_sheet1.cell(row=row_int, column=col_int).value = cleaned_int_status[element_index][r]
                                col_int += 1
                            else:
                                continue

                    # Clean BGP data which is received from file device_output_bgp.txt
                    re_pattern_neighbor_heading = re.compile(r"^Neigh")
                    re_pattern_neighbor_state = re.compile(r"(^([0-9]+\.))")
                    for k in range(len(bgp_summary_output)):
                        re_match_neighbor_heading = re_pattern_neighbor_heading.finditer(bgp_summary_output[k])
                        re_match_neighbor_state = re_pattern_neighbor_state.finditer(bgp_summary_output[k])

                        if re_match_neighbor_heading:
                            for neighbor_heading in re_match_neighbor_heading:
                                cleaned_bgp_summary.append(bgp_summary_output[k].rstrip("\n").split(" "))

                        if re_match_neighbor_state:
                            for neighbor_state in re_match_neighbor_state:
                                cleaned_bgp_summary.append(bgp_summary_output[k].rstrip("\n").split(" "))

                    # After cleaning BGP output we got list fom string and that is stored in already defined list so iterate
                    # across that list to fill in Excel sheet after checking it is null character or not
                    row_B = 0
                    for B in range(len(cleaned_bgp_summary)):
                        row_B += 1
                        col_B = 1

                        for C in range(len(cleaned_bgp_summary[B])):
                            # Check string/element is null or not if not null only after that it will add  in Excel and
                            # increase column number

                            if cleaned_bgp_summary[B][C] != "":
                                precheck_sheet2.cell(row=row_B, column=col_B).value = cleaned_bgp_summary[B][C]
                                col_B += 1


                    # Work for BFD neighbor
                    bfd_neighbor_only = []
                    re_pattern_bfd_heading = re.compile(r"^([a-z]|[A-Z])+")
                    re_pattern_bfd_neighbor_ip = re.compile(r"^([0-9]+\.)")   # This is sufficient to match for output
                    for bfd_neighbor_index in range(len(bfd_neighbor_output)):
                        re_match_bfd_heading =  re_pattern_bfd_heading.finditer(bfd_neighbor_output[bfd_neighbor_index])
                        re_match_bfd_neighbor_ip = re_pattern_bfd_neighbor_ip.finditer(bfd_neighbor_output[bfd_neighbor_index])
                        for bfd_heading_match in re_match_bfd_heading:
                            bfd_neighbor_only.append(bfd_neighbor_output[bfd_neighbor_index].rstrip("\n").split())
                        for bfd_neighbor_ip in re_match_bfd_neighbor_ip:
                            bfd_neighbor_only.append(bfd_neighbor_output[bfd_neighbor_index].rstrip("\n").split())

                    bfd_sheet_row =1
                    for bfd_row_index in range(len(bfd_neighbor_only)):
                        bfd_sheet_col = 1
                        for bfd_col_index in range(len(bfd_neighbor_only[bfd_row_index])):

                            if bfd_neighbor_only[bfd_row_index][bfd_col_index] != "":
                                precheck_sheet3.cell(row=bfd_sheet_row, column=bfd_sheet_col).value = bfd_neighbor_only[bfd_row_index][bfd_col_index]
                                bfd_sheet_col +=1
                        bfd_sheet_row +=1

                    # Apply formatting in precheck sheets headers
                    for heading_col_precheck_int in range(1, precheck_sheet1.max_column + 1):
                        precheck_sheet1.cell(row=1, column=heading_col_precheck_int).style = heading_style_template_precheck

                    for heading_col_precheck_bgp in range(1, precheck_sheet2.max_column + 1):
                        precheck_sheet2.cell(row=1, column=heading_col_precheck_bgp).style = heading_style_template_precheck
                    for heading_col_precheck_bfd in range(1, precheck_sheet3.max_column+1):
                        precheck_sheet3.cell(row=1, column=heading_col_precheck_bfd).style = heading_style_template_precheck

                    # Now apply formatting to cells
                    for cell_row_precheck_int in range(2, precheck_sheet1.max_row+1):
                        for cell_col_precheck_int in range(1, precheck_sheet1.max_column+1):
                            precheck_sheet1.cell(row=cell_row_precheck_int, column=cell_col_precheck_int).style = cell_style_template_precheck

                    for cell_row_precheck_bgp in range(2, precheck_sheet2.max_row+1):
                        for cell_col_precheck_bgp in range(1, precheck_sheet2.max_column+1):
                            precheck_sheet2.cell(row=cell_row_precheck_bgp, column=cell_col_precheck_bgp).style = cell_style_template_precheck

                    for cell_row_precheck_bfd in range (2, precheck_sheet3.max_row+1):
                        for cell_col_precheck_bfd in range(1, precheck_sheet3.max_column+1):
                            precheck_sheet3.cell(row=cell_row_precheck_bfd, column= cell_col_precheck_bfd).style = cell_style_template_precheck

                    wb_pre.save(r"Device_Output_Files\{device_name}.xlsx".format(device_name = i))
                    wb_pre.close()
    except Exception as Error:
        print(Error)

elif purpose.lower() == 'post':
    for device in device_list:
        int_status_post_clean = []
        int_status_cleaned_all = []
        bgp_status_post_cleaned = []
        with open(r"Device_Output_Files\device_output_int_post_{}.txt".format(device), mode = "w") as int_post:
        # Establish SSH Connection with device
            device_ssh = netmiko.ConnectHandler(host=device, username=username, password=password,
                                                device_type="cisco_nxos")
            print(device_ssh.find_prompt())
            int_status_post = device_ssh.send_command("show interface status")
            bgp_status_post = device_ssh.send_command("show ip bgp summary")
            bfd_status_post = device_ssh.send_command("show bfd neighbor")
            device_ssh.disconnect()

        # Open file to write interface and BGP output of device
            with open(r"Device_Output_Files\device_output_int_post_{}.txt".format(device), mode="w") as int_post_file:
                int_post_file.write(int_status_post)
            with open(r"Device_Output_Files\device_output_bgp_post_{}.txt".format(device), mode="w") as bgp_post_file:
                bgp_post_file.write(bgp_status_post)
            with open(r"Device_Output_Files\device_output_bfd_post_{}.txt".format(device), mode="w") as bfd_post_file:
                bfd_post_file.write(bfd_status_post)

        # Open device output files to store in variable
        with open(r"Device_Output_Files\device_output_int_post_{}.txt".format(device)) as device_int_post:
            int_status_post_all = device_int_post.readlines()
        with open(r"Device_Output_Files\device_output_bgp_post_{}.txt".format(device)) as device_bgp_post:
            bgp_status_post_all = device_bgp_post.readlines()
        with open(r"Device_Output_Files\device_output_bfd_post_{}.txt".format(device)) as device_bfd_post:
            bfd_status_post_all = device_bfd_post.readlines()

        # All those above values are in list now it's time to clean data and write in Excel
        # Separate elements of file to store in new list and put in Excel sheet
        row_int = 0
        col_int = 0
        wb_post = openpyxl.load_workbook(r"Device_Output_Files\{device_name}.xlsx".format(device_name = device ))
        postcheck_sheet1 = wb_post.create_sheet("Interface_postcheck")
        postcheck_sheet2 = wb_post.create_sheet(title="BGP_Summary_postcheck")
        postcheck_sheet3 = wb_post.create_sheet(title="BFD_neighbor_postcheck")


        # define pattern and based on that separate from interface status output and enter in Excel
        re_pattern_any_char = re.compile(r"^(([a-z]|[A-Z])+)")  # To match interface number and legends
        for fetch_index in range(len(int_status_post_all)):
            re_match_any_char = re_pattern_any_char.finditer(int_status_post_all[fetch_index])
            if re_match_any_char:
                for int_found in re_match_any_char:
                    int_status_post_clean.append(int_status_post_all[fetch_index].rstrip("\n").split(" "))

        # First write interface status in Excel which we got after splitting as per match condition
        for int_row_index in range(len(int_status_post_clean)):
            row_int += 1
            col_int = 1
            for elements_no in range(len(int_status_post_clean[int_row_index])):
                if int_status_post_clean[int_row_index][elements_no] != "":  # If elements is not null then add in Excel
                    postcheck_sheet1.cell(row=row_int, column=col_int).value = int_status_post_clean[int_row_index][elements_no]
                    col_int += 1
                else:
                    continue

        # Clean BGP output data only for neighbors as per pattern
        re_pattern_bgp_neighbor_heading = re.compile(r"^(Neigh)")
        re_pattern_bgp_neighbor_state = re.compile(r"^(([0-9]+)\.)")
        for bgp_output_index in range(len(bgp_status_post_all)):
            re_match_neighbor_heading = re_pattern_bgp_neighbor_heading.finditer(bgp_status_post_all[bgp_output_index])
            re_match_neighbor_state = re_pattern_bgp_neighbor_state.finditer(bgp_status_post_all[bgp_output_index])

            if re_match_neighbor_heading:  # If match found then separate and add in list
                for neighbor_heading in re_match_neighbor_heading:
                    bgp_status_post_cleaned.append(bgp_status_post_all[bgp_output_index].rstrip("\n").split())

            if re_match_neighbor_state:  # If match found then separate and add in list
                for neighbor_state in re_match_neighbor_state:
                    bgp_status_post_cleaned.append(bgp_status_post_all[bgp_output_index].rstrip("\n").split())

        # Iterate in cleaned bgp output list to add values in sheet
        row_bgp = 0
        col_bgp = 0
        for element_index in range(len(bgp_status_post_cleaned)):
            row_bgp += 1
            col_bgp = 1
            for element in range(len(bgp_status_post_cleaned[element_index])):
                postcheck_sheet2.cell(row=row_bgp, column=col_bgp).value = bgp_status_post_cleaned[element_index][element]
                col_bgp += 1


        # Work for BFD to match, split, and enter in postchek sheet
        re_pattern_bfd_neighbor_heading = re.compile(r"^([a-z]|[A-Z])+")
        re_pattern_bfd_neighbor_ip = re.compile(r"^([0-9]+\.)")
        bfd_neighbor_only_post = []
        for bfd_output_length in range(len(bfd_status_post_all)):
            re_match_bfd_neighbor_heading = re_pattern_bfd_neighbor_heading.finditer(bfd_status_post_all[bfd_output_length])
            re_match_bfd_neighbor_ip = re_pattern_bfd_neighbor_ip.finditer(bfd_status_post_all[bfd_output_length])
            for bfd_neighbor_heading_match in re_match_bfd_neighbor_heading:
                bfd_neighbor_only_post.append(bfd_status_post_all[bfd_output_length].rstrip("\n").split())
            for bfd_neighbor_ip_match in re_match_bfd_neighbor_ip:
                bfd_neighbor_only_post.append(bfd_status_post_all[bfd_output_length].rstrip("\n").split())


        bfd_row_post = 1
        for bfd_neighbor_line in range(len(bfd_neighbor_only_post)):
            bfd_column_post = 1
            for bfd_neighbor_line_element in range(len(bfd_neighbor_only_post[bfd_neighbor_line])):

                if bfd_neighbor_only_post[bfd_neighbor_line][bfd_neighbor_line_element] == "": # It is not working but kept
                    continue
                else:
                    postcheck_sheet3.cell(row=bfd_row_post, column=bfd_column_post).value = bfd_neighbor_only_post[bfd_neighbor_line][bfd_neighbor_line_element]
                    bfd_column_post +=1
            bfd_row_post +=1

        # Till now data available in Excel sheet. Formatting and comparison left, before proceeding further save file.
        # wb_post.save("{device_name}.xlsx".format(device_name = device))

        '''
        Vlokup formula for interface status = "=VLOOKUP(A2,Interfaces_precheck!A:G,3,0)"
        Vlookup formula for BGP staus = "=VLOOKUP(A2,BGP_Summary_Precheck!A:J,9,0)"
        Vlookup for BGP pfx = "=VLOOKUP(A2,BGP_Summary_Precheck!A:K,10,0)"
        Vlookup for BFD state = "=VLOOKUP(A3,BFD_neighbor_precheck!A:H,6,0)"
        Condition check formula for inteface = "=IF(C2=I2,"OK","Changed")"
        Condition check for BGP state = "=IF(I2=K2,"OK", "Changed")"
        Condition check for BGP pfx = "=IF(J2=M2,"OK", "Changed")"
        Condition check for BFD state = "=IF(F2=I2, "OK","Changed")"
        '''
        # Load precheck workbook
        total_sheets = wb_post.sheetnames
        precheck_sheet1_int = wb_post[total_sheets[0]]
        precheck_sheet2_bgp = wb_post[total_sheets[1]]
        precheck_sheet3_bfd = wb_post[total_sheets[2]]
        col_for_vlookup_result_int = postcheck_sheet1.max_column # If i put inside loop then after first iteration column added and in next have one more column and subsequently column increases with +1
        postcheck_sheet1.cell(row=1, column= postcheck_sheet1.max_column).value = "Status in Precheck"
        postcheck_sheet1.cell(row=1, column=postcheck_sheet1.max_column + 1).value = "Compare result"
        col_for_compare_result_int = postcheck_sheet1.max_column

        # Check status of interface and put it in postcheck sheet of interface
        for row_vlookup_compare_int in range(2, postcheck_sheet1.max_row+1):
            postcheck_sheet1.cell(row=row_vlookup_compare_int, column= col_for_vlookup_result_int).value = "=vlookup(A{row_no}, {sheetname_precheck}!A:{max_column},3,0)".format(
                row_no =row_vlookup_compare_int, sheetname_precheck= total_sheets[0], max_column = openpyxl.utils.get_column_letter(precheck_sheet1_int.max_column))
            postcheck_sheet1.cell(row=row_vlookup_compare_int, column= col_for_compare_result_int).value = '=if(C{cell_row} = {col_result}{cell_row}, "OK", "Changed")'.format(
                cell_row=row_vlookup_compare_int, col_result = openpyxl.utils.get_column_letter(col_for_vlookup_result_int))

        # Check status of BGP and put it in postcheck sheet of BGP
        postcheck_sheet2.cell(row=1, column=postcheck_sheet2.max_column+1).value = "BGP state in Precheck"
        col_for_vlookup_bgp_state = postcheck_sheet2.max_column
        postcheck_sheet2.cell(row=1, column=postcheck_sheet2.max_column + 1).value = "BGP neighbor pfx in Precheck"
        col_for_vlookup_bgp_pfx = postcheck_sheet2.max_column
        postcheck_sheet2.cell(row=1, column=postcheck_sheet2.max_column+1).value = "Compare result of State"
        col_for_compare_result_bgp_state = postcheck_sheet2.max_column
        postcheck_sheet2.cell(row=1, column=postcheck_sheet2.max_column + 1).value = "Compare result of pfx"
        col_for_compare_result_bgp_pfx = postcheck_sheet2.max_column

        for row_vlookup_bgp in range(2, postcheck_sheet2.max_row+1):

            postcheck_sheet2.cell(row=row_vlookup_bgp, column= col_for_vlookup_bgp_state).value = "=vlookup(A{neighbor_ip}, {name_bgp_sheet_precheck}!A:{max_column_in_bgp_precheck},9,0)".format(
                neighbor_ip=row_vlookup_bgp, name_bgp_sheet_precheck=total_sheets[1],
                max_column_in_bgp_precheck=openpyxl.utils.get_column_letter(precheck_sheet2_bgp.max_column))

            postcheck_sheet2.cell(row=row_vlookup_bgp,
                                  column=col_for_vlookup_bgp_pfx).value = "=vlookup(A{neighbor_ip}, {name_bgp_sheet_precheck}!A:{max_column_in_bgp_precheck},10,0)".format(
                neighbor_ip=row_vlookup_bgp, name_bgp_sheet_precheck=total_sheets[1],
                max_column_in_bgp_precheck=openpyxl.utils.get_column_letter(precheck_sheet2_bgp.max_column))

            postcheck_sheet2.cell(row=row_vlookup_bgp, column=col_for_compare_result_bgp_state).value = '=if(I{row_index} = {col_result}{row_index}, "OK","Changed")'.format(
                row_index=row_vlookup_bgp, col_result= openpyxl.utils.get_column_letter(col_for_vlookup_bgp_state))

            postcheck_sheet2.cell(row=row_vlookup_bgp, column=col_for_compare_result_bgp_pfx).value = '=if(J{row_index} = {col_result}{row_index}, "OK","Changed")'.format(
                row_index=row_vlookup_bgp, col_result= openpyxl.utils.get_column_letter(col_for_vlookup_bgp_pfx))

        # Apply vlookup and condition for BFD
        postcheck_sheet3.cell(row=1, column=postcheck_sheet3.max_column+1).value = "State in precheck"
        bfd_vlookup_cell_column = postcheck_sheet3.max_column

        postcheck_sheet3.cell(row=1, column=postcheck_sheet3.max_column+1).value = "Compare result"
        bfd_compare_cell_column = postcheck_sheet3.max_column

        for row_vlookup_bfd in range(2, postcheck_sheet3.max_row+1):
            postcheck_sheet3.cell(row=row_vlookup_bfd, column=bfd_vlookup_cell_column).value = "=VLOOKUP(A{neighbor_bfd_ip_row},{sheetname_bfd}!A:{max_column_in_precheck},6,0)".format(neighbor_bfd_ip_row=row_vlookup_bfd,
                                                                                                                                                                                    sheetname_bfd=total_sheets[2], max_column_in_precheck = openpyxl.utils.get_column_letter(postcheck_sheet3.max_column))

            postcheck_sheet3.cell(row=row_vlookup_bfd, column=bfd_compare_cell_column).value = '=IF(F{row}=I{row}, "OK","Changed")'.format(row=row_vlookup_bfd)


        # Now apply formatting to headings of sheets
        for col_format_heading_int in range(1, postcheck_sheet1.max_column+1):
            postcheck_sheet1.cell(row=1, column=col_format_heading_int).style = heading_style_template_postcheck

        for col_format_heading_bgp in range(1, postcheck_sheet2.max_column+1):
            postcheck_sheet2.cell(row=1, column=col_format_heading_bgp).style = heading_style_template_postcheck

        for col_format_heading_bfd in range(1, postcheck_sheet3.max_column+1):
            postcheck_sheet3.cell(row=1, column=col_format_heading_bfd).style = heading_style_template_postcheck

        # Now apply formatting to cells of sheet
        for row_format_cell_int in range(2, postcheck_sheet1.max_row+1):
            for col_format_cell_int in range(1, postcheck_sheet1.max_column+1):
                postcheck_sheet1.cell(row=row_format_cell_int, column=col_format_cell_int).style = cell_style_template_postcheck

        for row_format_cell_bgp in range(2, postcheck_sheet2.max_row+1):
            for col_format_cell_bgp in range(1, postcheck_sheet2.max_column+1):
                postcheck_sheet2.cell(row=row_format_cell_bgp, column=col_format_cell_bgp).style = cell_style_template_postcheck

        for row_format_cell_bfd in range(2, postcheck_sheet3.max_row+1):
            for col_format_cell_bfd in range(1, postcheck_sheet3.max_column+1):
                postcheck_sheet3.cell(row=row_format_cell_bfd, column=col_format_cell_bfd).style = cell_style_template_postcheck

        #   Now apply conditional formatting to highlight cell which are changed
        # cell_red_fill = openpyxl.styles.PatternFill(patternType="solid", start_color="00FF8080", end_color="00FF8080")
        # cell_red_fil_dxf = openpyxl.styles.differential.DifferentialStyle(fill=cell_red_fill)
        # cell_red_fill_condition_rule = openpyxl.formatting.rule.Rule(dxf=cell_red_fil_dxf, type="containsText", operator="containsText", text="Changed")
        # postcheck_sheet1.conditional_formatting.add("{column}2:{column}{row_end}".format(column=postcheck_sheet1.max_column,
        #                                                                                  row_end=postcheck_sheet1.max_row),
        #                                             cfRule=cell_red_fill_condition_rule)
        # postcheck_sheet2.conditional_formatting.add("{max_column}2:{max_column}{row_end}".format(max_column=postcheck_sheet2.max_column, row_end=postcheck_sheet2.max_row),
        #                                             cfRule= cell_red_fill_condition_rule)


        wb_post.save("{filename}.xlsx".format(filename = device))
        wb_post.close()


else:
    print("Wrong purpose")
