import openpyxl
import openpyxl.utils
import openpyxl.formatting
import openpyxl.formula
import openpyxl.styles
from openpyxl.styles.differential import DifferentialStyle
import openpyxl.formatting.rule
import netmiko


# For pre check device output
with open("int_status_pre.txt") as f:
    status = f.readlines()

# For post check device output
with open("int_status_post.txt") as f:
    status2 = f.readlines()

wb = openpyxl.Workbook()  # open workbook
ws_sheet3 = wb.active
ws_sheet1 = wb.create_sheet("precheck")  # Create new worksheet where to save processed data
ws_sheet2 = wb.create_sheet("postcheck")  # Print data after processing first level
ws_sheet3.title = "summary"

# Define two variable globallyn to store data after processing SSH output of netmiko

new_status = []
row = 0
r = 0

# Loop to remove new line and space character
for i in range(len(status)):
    new_status.append(status[i].rstrip("\n").split(" "))

    # After splitting string of previous netmiko list become string and each element become after splitting
    # Running new loop to append only those content which do not have "" it will remove all elements which are
    # introduced after splitting content.

    row += 1
    col = 1
    for a in range(len(new_status[i])):
        if new_status[i][a] != "":
            ws_sheet1.cell(row=row, column=col).value = new_status[i][a]
            col = col + 1
        else:
            continue

# Delete 1 and second row because it contain only ---------- for Nexus
ws_sheet1.delete_rows(1)
ws_sheet1.delete_rows(2)

# Now load same data as a postcheck and write in postcheck sheet
postcheck_trim = []
postcheck_row = 0

for i in range(len(status2)):
    postcheck_trim.append(status2[i].rstrip("\n").split(" "))  # After trimming each index element of status become list
    postcheck_col = 1
    postcheck_row += 1
    for j in range(len(postcheck_trim[i])):
        if postcheck_trim[i][j] != "":
            ws_sheet2.cell(row=postcheck_row, column=postcheck_col).value = postcheck_trim[i][j]
            postcheck_col += 1
        else:
            continue

ws_sheet2.delete_rows(1)
ws_sheet2.delete_rows(2)

# Copy all data from sheet2 to summary sheet
ws_sheet3 = wb.copy_worksheet(ws_sheet2)
ws_sheet3.title = "Summary"

# Remove sheet from beginning because first null sheet created. We can modify as well to no need this line.
sheet_list = wb.sheetnames
del wb[sheet_list[0]]
sheets = wb.sheetnames

# Add 2 column to record status of pre check and compare status in summary sheet
ws_sheet3.cell(row=1, column=ws_sheet3.max_column + 1).value = "Status in Pre check"
ws_sheet3.cell(row=1, column=ws_sheet3.max_column + 1).value = "Status changed/Not-Changed"

# Perform VLOOOKUP to fetch data from precheck
# =VLOOKUP(A2,precheck!A:G,3,0)
# =IF(C2=J2,"Not Changed","Changed")

# Apply VLOOKUP to fetch data from pre check sheet for interface/tunnel status
for i in range(2, ws_sheet3.max_row + 1):
    for j in range(ws_sheet3.max_column - 1, ws_sheet3.max_column):
        char = openpyxl.utils.get_column_letter(1)
        ws_sheet3.cell(row=i, column=j).value = "=vlookup({},precheck!A:{},3,0)".format((char + str(i)),
                                                                                (openpyxl.utils.get_column_letter
                                                                                            (ws_sheet1.max_row)))

# Apply if formula to compare between both status in summary sheet
for i in range(2, ws_sheet3.max_row + 1):
    for j in range(ws_sheet3.max_column, ws_sheet3.max_column + 1):
        ws_sheet3.cell(row=i, column=j).value = '=if({}={}, "Not-Changed", "Changed")'.format(("C" + str(i)),
                                                                                (openpyxl.utils.get_column_letter(
                                                                                 ws_sheet3.max_column - 1) + str(i)))

# Apply formatting to header of each sheet

header_font = openpyxl.styles.Font(sz=15, bold=True, color="00FF6600")

for i in range(1, 4):
    if i == 1:
        for r in range(1, 2):
            for c in range(1, ws_sheet1.max_column + 1):
                ws_sheet1.cell(row=r, column=c).font = header_font

    elif i == 2:
        for r in range(1, 2):
            for c in range(1, ws_sheet2.max_column + 1):
                ws_sheet2.cell(row=r, column=c).font = header_font
    elif i == 3:
        for r in range(1, 2):
            for c in range(1, ws_sheet3.max_column + 1):
                ws_sheet3.cell(row=r, column=c).font = header_font

# Now apply formatting to cells we can use named style as well

cell_border = openpyxl.styles.Border(left=openpyxl.styles.Side(border_style="thin", color="00000000"),
                                     right=openpyxl.styles.Side(border_style="thin", color="00000000"),
                                     top=openpyxl.styles.Side(border_style="thin", color="00000000"),
                                     bottom=openpyxl.styles.Side(border_style="thin", color="00000000"))

cell_font = openpyxl.styles.Font(sz=11)
cell_alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
cell_font_change = openpyxl.styles.Font(color="00FF0000")

for i in range(1, 4):
    if i == 1:
        for R in range(2, ws_sheet1.max_row + 2):  # Row is from 2 onwards max row +2 because we want border similar
                                                   # to cell but fornt styles which was earlier defined.
            for C in range(1, ws_sheet1.max_column + 1):
                ws_sheet1.cell(row=R, column=C).font = cell_font
                ws_sheet1.cell(row=R - 1, column=C).border = cell_border
                ws_sheet1.cell(row=R, column=C).alignment = cell_alignment
    if i == 2:
        for R in range(2, ws_sheet2.max_row + 2):
            for C in range(1, ws_sheet2.max_column + 1):
                ws_sheet2.cell(row=R, column=C).font = cell_font
                ws_sheet2.cell(row=R - 1, column=C).border = cell_border
                ws_sheet2.cell(row=R, column=C).alignment = cell_alignment
    if i == 3:
        for R in range(2, ws_sheet3.max_row + 2):
            for C in range(1, ws_sheet3.max_column + 1):
                ws_sheet3.cell(row=R, column=C).font = cell_font
                ws_sheet3.cell(row=R - 1, column=C).border = cell_border
                ws_sheet3.cell(row=R, column=C).alignment = cell_alignment


cell_fill_red = openpyxl.styles.PatternFill(start_color="00FF8080", end_color="00FF8080", fill_type="solid")
cell_dxf = DifferentialStyle(fill=cell_fill_red)

cell_rule = openpyxl.formatting.rule.Rule(type="beginsWith", operator="beginsWith", text="Changed", dxf = cell_dxf)

ws_sheet3.conditional_formatting.add("K2:K150", cell_rule)

wb.save("int_status_Compare_Result.xlsx")
wb.close()
